from __future__ import annotations

import re
from collections.abc import Iterable
from datetime import datetime
from functools import lru_cache
from io import BytesIO
from statistics import mean
from typing import Any

import openpyxl
import requests

import valuation_db

TIMEOUT_SEC = 30
REQUEST_RETRY_COUNT = 3
LAND_COMPLEX_API_URL = "https://api.kbland.kr/land-complex"
LAND_PRICE_API_URL = "https://api.kbland.kr/land-price"

COMMON_HEADERS = {
    "accept": "application/json, text/plain, */*",
    "origin": "https://kbland.kr",
    "referer": "https://kbland.kr/",
    "user-agent": "Mozilla/5.0",
}

KB_SIDO_ALIASES = {
    "서울": "서울시",
    "서울시": "서울시",
    "서울특별시": "서울시",
    "부산": "부산시",
    "부산시": "부산시",
    "부산광역시": "부산시",
    "대구": "대구시",
    "대구시": "대구시",
    "대구광역시": "대구시",
    "인천": "인천시",
    "인천시": "인천시",
    "인천광역시": "인천시",
    "광주": "광주시",
    "광주시": "광주시",
    "광주광역시": "광주시",
    "대전": "대전시",
    "대전시": "대전시",
    "대전광역시": "대전시",
    "울산": "울산시",
    "울산시": "울산시",
    "울산광역시": "울산시",
    "세종": "세종시",
    "세종시": "세종시",
    "세종특별자치시": "세종시",
    "경기": "경기도",
    "경기도": "경기도",
    "강원": "강원도",
    "강원도": "강원도",
    "강원특별자치도": "강원도",
    "충북": "충청북도",
    "충청북도": "충청북도",
    "충남": "충청남도",
    "충청남도": "충청남도",
    "전북": "전라북도",
    "전라북도": "전라북도",
    "전북특별자치도": "전라북도",
    "전남": "전라남도",
    "전라남도": "전라남도",
    "경북": "경상북도",
    "경상북도": "경상북도",
    "경남": "경상남도",
    "경상남도": "경상남도",
    "제주": "제주도",
    "제주도": "제주도",
    "제주특별자치도": "제주도",
}

MONTH_KEY_PATTERN = re.compile(r"^(20\d{2})(\d{2})$")
DATE_TOKEN_CANDIDATES = (
    "기준년월",
    "년월",
    "조사년월",
    "기준월",
    "시세년월",
    "기준일",
    "조사일자",
    "날짜",
)

SALE_GENERAL_KEYS = (
    "매매일반거래가",
    "매매일반가",
    "매매일반평균가",
    "일반평균가",
    "일반가",
)
JEONSE_GENERAL_KEYS = (
    "전세일반거래가",
    "전세일반가",
    "전세일반평균가",
)
SALE_UPPER_KEYS = ("매매상위평균가", "상위평균가")
SALE_LOWER_KEYS = ("매매하위평균가", "하위평균가")
JEONSE_UPPER_KEYS = ("전세상위평균가",)
JEONSE_LOWER_KEYS = ("전세하위평균가",)

SESSION = requests.Session()
SESSION.headers.update(COMMON_HEADERS)


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _normalize_int(value: Any) -> int | None:
    text = _clean_text(value)
    if not text:
        return None
    digits = re.sub(r"[^\d-]", "", text)
    if not digits or digits == "-":
        return None
    return int(digits)


def _normalize_float(value: Any) -> float | None:
    text = _clean_text(value)
    if not text:
        return None
    digits = re.sub(r"[^\d.-]", "", text)
    if not digits or digits in {"-", ".", "-."}:
        return None
    return float(digits)


def _normalize_month_key(value: Any) -> str | None:
    text = _clean_text(value)
    if not text:
        return None

    digits = re.sub(r"[^\d]", "", text)
    if len(digits) == 6 and MONTH_KEY_PATTERN.match(digits):
        return digits
    if len(digits) >= 8 and MONTH_KEY_PATTERN.match(digits[:6]):
        return digits[:6]

    match = re.search(r"(20\d{2})[.\-/년 ]?(\d{1,2})", text)
    if match:
        year = match.group(1)
        month = int(match.group(2))
        return f"{year}{month:02d}"

    match = re.search(r"(\d{2})[.\-/년 ]?(\d{1,2})", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        return f"20{year:02d}{month:02d}"

    return None


def _month_label(month_key: str | None) -> str | None:
    if not month_key or not MONTH_KEY_PATTERN.match(month_key):
        return None
    return f"{month_key[:4]}-{month_key[4:6]}"


def _request_json(base_url: str, path: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
    last_error: Exception | None = None
    for _attempt in range(REQUEST_RETRY_COUNT):
        try:
            response = SESSION.get(f"{base_url}{path}", params=params, timeout=TIMEOUT_SEC)
            response.raise_for_status()
            payload = response.json()
            header = payload.get("dataHeader", {})
            if header.get("resultCode") != "10000":
                raise RuntimeError(f"KB API 오류: {header}")
            body = payload.get("dataBody", {})
            result_code = body.get("resultCode")
            if result_code not in (None, 11000, "11000", 31210, "31210", 33210, "33210"):
                raise RuntimeError(f"KB API dataBody 오류: {body}")
            return body
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"KB API 요청 실패: {base_url}{path}") from last_error


def _download_bytes(base_url: str, path: str, params: dict[str, Any]) -> bytes:
    last_error: Exception | None = None
    for _attempt in range(REQUEST_RETRY_COUNT):
        try:
            response = SESSION.get(f"{base_url}{path}", params=params, timeout=TIMEOUT_SEC)
            response.raise_for_status()
            return response.content
        except Exception as exc:
            last_error = exc
    raise RuntimeError(f"KB 파일 다운로드 실패: {base_url}{path}") from last_error


@lru_cache(maxsize=1)
def _get_sido_list() -> tuple[dict[str, Any], ...]:
    body = _request_json(LAND_COMPLEX_API_URL, "/map/siDoAreaNameList")
    return tuple(body.get("data", []))


@lru_cache(maxsize=64)
def _get_sigungu_list(sido_name: str) -> tuple[dict[str, Any], ...]:
    body = _request_json(
        LAND_COMPLEX_API_URL,
        "/map/siGunGuAreaNameList",
        params={"시도명": sido_name},
    )
    return tuple(body.get("data", []))


@lru_cache(maxsize=256)
def _get_dong_list(sido_name: str, sigungu_name: str) -> tuple[dict[str, Any], ...]:
    body = _request_json(
        LAND_COMPLEX_API_URL,
        "/map/stutDongAreaNameList",
        params={"시도명": sido_name, "시군구명": sigungu_name},
    )
    return tuple(body.get("data", []))


@lru_cache(maxsize=1024)
def _get_complexes_by_dong_code(dong_code: str) -> tuple[dict[str, Any], ...]:
    body = _request_json(
        LAND_COMPLEX_API_URL,
        "/complexComm/hscmList",
        params={"법정동코드": dong_code},
    )
    return tuple(body.get("data", []))


@lru_cache(maxsize=2048)
def _get_complex_types_remote(complex_id: int) -> tuple[dict[str, Any], ...]:
    body = _request_json(
        LAND_COMPLEX_API_URL,
        "/complex/typInfo",
        params={"단지기본일련번호": complex_id},
    )
    return tuple(body.get("data", []))


@lru_cache(maxsize=2048)
def _get_complex_main(complex_id: int) -> dict[str, Any]:
    body = _request_json(
        LAND_COMPLEX_API_URL,
        "/complex/main",
        params={"단지기본일련번호": complex_id},
    )
    data = body.get("data")
    if isinstance(data, dict):
        return data
    return {}


def _normalize_region_name(region_name: str) -> str:
    region_name = _clean_text(region_name)
    if not region_name:
        raise ValueError("지역명이 비어 있습니다.")
    tokens = region_name.split(" ")
    if tokens and tokens[0] in KB_SIDO_ALIASES:
        tokens[0] = KB_SIDO_ALIASES[tokens[0]]
    return " ".join(tokens)


def _resolve_sigungu_from_short_name(short_name: str) -> tuple[str, str]:
    matches: list[tuple[str, str]] = []
    for sido in _get_sido_list():
        kb_sido_name = _clean_text(sido.get("시도명"))
        for sigungu in _get_sigungu_list(kb_sido_name):
            sigungu_name = _clean_text(sigungu.get("시군구명"))
            if sigungu_name == short_name:
                matches.append((kb_sido_name, sigungu_name))
    if not matches:
        raise ValueError(f"KB 지역 스코프를 찾을 수 없습니다: {short_name}")
    if len(matches) > 1:
        raise ValueError(f"동일한 시군구명이 여러 시도에 존재합니다: {short_name}")
    return matches[0]


def _resolve_region_scope(region_name: str) -> tuple[str, str]:
    normalized = _normalize_region_name(region_name)
    tokens = normalized.split(" ")
    sido_names = {_clean_text(item.get("시도명")) for item in _get_sido_list()}
    if tokens[0] in sido_names:
        if len(tokens) == 1:
            raise ValueError(f"시도만으로는 단지 캐시를 갱신할 수 없습니다: {region_name}")
        return tokens[0], " ".join(tokens[1:])
    return _resolve_sigungu_from_short_name(normalized)


def _iter_region_complexes(region_name: str) -> list[dict[str, Any]]:
    sido_name, sigungu_name = _resolve_region_scope(region_name)
    complexes: dict[int, dict[str, Any]] = {}
    for dong in _get_dong_list(sido_name, sigungu_name):
        dong_code = _clean_text(dong.get("법정동코드"))
        dong_name = _clean_text(dong.get("법정동명") or dong.get("법정동읍면동명"))
        if not dong_code:
            continue
        for complex_info in _get_complexes_by_dong_code(dong_code):
            if _clean_text(complex_info.get("매물종별구분")) != "01":
                continue
            complex_id = complex_info.get("단지기본일련번호")
            if not isinstance(complex_id, int):
                continue
            merged = dict(complex_info)
            merged["_sido_name"] = sido_name
            merged["_sigungu_name"] = sigungu_name
            merged["_dong_name"] = dong_name
            merged["_dong_code"] = dong_code
            complexes[complex_id] = merged
    return list(complexes.values())


def _normalize_complex_row(complex_info: dict[str, Any]) -> dict[str, Any]:
    complex_id = int(complex_info["단지기본일련번호"])
    main_info = {}
    try:
        main_info = _get_complex_main(complex_id)
    except Exception:
        main_info = {}

    address = (
        _clean_text(main_info.get("주소"))
        or _clean_text(main_info.get("지번주소"))
        or _clean_text(complex_info.get("주소"))
    )
    households = (
        _normalize_int(main_info.get("총세대수"))
        or _normalize_int(complex_info.get("세대수"))
    )
    completion_year = _normalize_int(main_info.get("준공년수")) or _normalize_int(main_info.get("준공년도"))
    if completion_year and completion_year < 100:
        current_year = datetime.now().year
        completion_year = current_year - completion_year

    return {
        "complex_id": complex_id,
        "complex_name": _clean_text(complex_info.get("단지명")),
        "sido_name": _clean_text(complex_info.get("_sido_name")),
        "sigungu_name": _clean_text(complex_info.get("_sigungu_name")),
        "dong_name": _clean_text(complex_info.get("_dong_name")),
        "dong_code": _clean_text(complex_info.get("_dong_code")),
        "address": address,
        "households": households,
        "completion_year": completion_year,
        "entrance_type": _clean_text(main_info.get("현관구조")),
        "raw_json": {
            "complex_info": complex_info,
            "main_info": main_info,
        },
    }


def _normalize_type_row(type_info: dict[str, Any]) -> dict[str, Any] | None:
    area_id = type_info.get("면적일련번호")
    if not isinstance(area_id, int):
        return None
    return {
        "area_id": area_id,
        "type_name": _clean_text(type_info.get("타입명") or type_info.get("면적명")),
        "exclusive_area": _normalize_float(type_info.get("전용면적")),
        "exclusive_area_pyeong": _normalize_float(type_info.get("전용면적평")),
        "supply_area": _normalize_float(type_info.get("공급면적")),
        "supply_area_pyeong": _normalize_float(type_info.get("공급면적평")),
        "contract_area": _normalize_float(type_info.get("계약면적")),
        "contract_area_pyeong": _normalize_float(type_info.get("계약면적평")),
        "households": _normalize_int(type_info.get("세대수")),
        "room_count": _normalize_int(type_info.get("방수")),
        "bathroom_count": _normalize_int(type_info.get("욕실수")),
        "total_trade_count": _normalize_int(type_info.get("매매건수")),
        "total_jeonse_count": _normalize_int(type_info.get("전세건수")),
        "total_monthly_rent_count": _normalize_int(type_info.get("월세건수")),
        "raw_json": type_info,
    }


def search_cached_complexes(query: str, limit: int = 20) -> list[dict[str, Any]]:
    return valuation_db.search_complexes(query, limit=limit)


def sync_region_complex_cache(region_name: str, *, sync_types: bool = False) -> dict[str, Any]:
    complexes = _iter_region_complexes(region_name)
    normalized_rows = [_normalize_complex_row(item) for item in complexes]
    valuation_db.upsert_complexes(normalized_rows)

    synced_types = 0
    if sync_types:
        for row in normalized_rows:
            synced_types += len(ensure_complex_types_cached(row["complex_id"], force_refresh=True))

    return {
        "region_name": region_name,
        "complex_count": len(normalized_rows),
        "type_count": synced_types,
        "cached_total_complexes": valuation_db.count_cached_complexes(),
    }


def ensure_complex_types_cached(complex_id: int, *, force_refresh: bool = False) -> list[dict[str, Any]]:
    cached_rows = valuation_db.get_complex_types(complex_id)
    if cached_rows and not force_refresh:
        return cached_rows

    normalized_rows: list[dict[str, Any]] = []
    for type_info in _get_complex_types_remote(complex_id):
        normalized = _normalize_type_row(type_info)
        if normalized:
            normalized_rows.append(normalized)
    valuation_db.upsert_complex_types(complex_id, normalized_rows)
    return valuation_db.get_complex_types(complex_id)


def get_complex_type_options(complex_id: int, *, force_refresh: bool = False) -> list[dict[str, Any]]:
    rows = ensure_complex_types_cached(complex_id, force_refresh=force_refresh)
    options: list[dict[str, Any]] = []
    for row in rows:
        area_text = f"{row.get('exclusive_area') or '-'}㎡"
        area_pyeong = row.get("exclusive_area_pyeong")
        households = row.get("households")
        parts = [area_text]
        if area_pyeong:
            parts.append(f"{area_pyeong}평")
        if households:
            parts.append(f"{households}세대")
        options.append(
            {
                "area_id": row["area_id"],
                "exclusive_area": row.get("exclusive_area"),
                "exclusive_area_pyeong": row.get("exclusive_area_pyeong"),
                "label": " / ".join(parts),
                "type_name": row.get("type_name"),
            }
        )
    return options


def _history_years(lookback_years: int, as_of_month: str | None) -> list[int]:
    if as_of_month and MONTH_KEY_PATTERN.match(as_of_month):
        latest_year = int(as_of_month[:4])
    else:
        latest_year = datetime.now().year
    first_year = latest_year - max(lookback_years, 1)
    return list(range(first_year, latest_year + 1))


def _request_quote_history_json(complex_id: int, area_id: int, *, years: list[int]) -> list[dict[str, Any]]:
    params_candidates = [
        {
            "단지기본일련번호": complex_id,
            "면적일련번호": area_id,
            "기준년": ",".join(str(year) for year in years),
        },
        {
            "단지기본일련번호": complex_id,
            "면적기준일련번호": area_id,
            "기준년": ",".join(str(year) for year in years),
        },
    ]
    last_error: Exception | None = None
    for params in params_candidates:
        try:
            body = _request_json(LAND_PRICE_API_URL, "/price/WholQuotList", params=params)
            rows = _parse_quote_history_rows(body.get("data"))
            if rows:
                return rows
        except Exception as exc:
            last_error = exc
    if last_error:
        raise RuntimeError("KB 시세 이력 JSON 조회에 실패했습니다.") from last_error
    return []


def _download_quote_history_excel(complex_id: int, area_id: int) -> list[dict[str, Any]]:
    params_candidates = [
        {
            "단지기본일련번호": complex_id,
            "면적일련번호": area_id,
            "연결구분명": "일반",
        },
        {
            "단지기본일련번호": complex_id,
            "면적기준일련번호": area_id,
            "연결구분명": "일반",
        },
    ]
    last_error: Exception | None = None
    for params in params_candidates:
        try:
            content = _download_bytes(LAND_PRICE_API_URL, "/price/perMnPastPriceExcelDownload", params=params)
            rows = _parse_quote_history_excel(content)
            if rows:
                return rows
        except Exception as exc:
            last_error = exc
    if last_error:
        raise RuntimeError("KB 시세 이력 엑셀 다운로드에 실패했습니다.") from last_error
    return []


def _extract_first_value(payload: dict[str, Any], keys: Iterable[str]) -> int | None:
    for key in keys:
        if key in payload:
            value = _normalize_int(payload.get(key))
            if value is not None:
                return value
    return None


def _normalize_history_row(item: dict[str, Any]) -> dict[str, Any] | None:
    month_key = None
    for key in DATE_TOKEN_CANDIDATES:
        month_key = _normalize_month_key(item.get(key))
        if month_key:
            break

    if not month_key:
        for key, value in item.items():
            if "년월" in key or key.endswith("월"):
                month_key = _normalize_month_key(value)
                if month_key:
                    break

    sale_general_price = _extract_first_value(item, SALE_GENERAL_KEYS)
    jeonse_general_price = _extract_first_value(item, JEONSE_GENERAL_KEYS)
    sale_upper_price = _extract_first_value(item, SALE_UPPER_KEYS)
    sale_lower_price = _extract_first_value(item, SALE_LOWER_KEYS)
    jeonse_upper_price = _extract_first_value(item, JEONSE_UPPER_KEYS)
    jeonse_lower_price = _extract_first_value(item, JEONSE_LOWER_KEYS)

    if not month_key:
        return None
    if sale_general_price is None and jeonse_general_price is None:
        return None

    return {
        "month_key": month_key,
        "month_label": _month_label(month_key),
        "sale_general_price": sale_general_price,
        "jeonse_general_price": jeonse_general_price,
        "sale_upper_price": sale_upper_price,
        "sale_lower_price": sale_lower_price,
        "jeonse_upper_price": jeonse_upper_price,
        "jeonse_lower_price": jeonse_lower_price,
    }


def _parse_quote_history_rows(data: Any) -> list[dict[str, Any]]:
    if not isinstance(data, dict):
        return []
    sections = data.get("시세")
    if not isinstance(sections, list):
        return []

    rows_by_month: dict[str, dict[str, Any]] = {}
    for section in sections:
        if not isinstance(section, dict):
            continue
        items = section.get("items")
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            normalized = _normalize_history_row(item)
            if normalized:
                rows_by_month[normalized["month_key"]] = normalized

    return sorted(rows_by_month.values(), key=lambda row: row["month_key"], reverse=True)


def _parse_quote_history_excel(content: bytes) -> list[dict[str, Any]]:
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    header_row_index: int | None = None
    header_map: dict[str, int] = {}

    for row_index in range(1, min(worksheet.max_row, 20) + 1):
        values = [_clean_text(worksheet.cell(row=row_index, column=col_index).value) for col_index in range(1, worksheet.max_column + 1)]
        if not any(values):
            continue
        local_header_map = {value: idx for idx, value in enumerate(values) if value}
        if any(candidate in local_header_map for candidate in ("월", "년월", "기준년월")) and any(
            candidate in local_header_map for candidate in SALE_GENERAL_KEYS
        ):
            header_row_index = row_index
            header_map = local_header_map
            break

    if header_row_index is None:
        return []

    month_col = None
    for key in ("월", "년월", "기준년월"):
        if key in header_map:
            month_col = header_map[key] + 1
            break
    if month_col is None:
        return []

    def read_value(row_number: int, keys: Iterable[str]) -> Any:
        for key in keys:
            if key in header_map:
                return worksheet.cell(row=row_number, column=header_map[key] + 1).value
        return None

    rows: list[dict[str, Any]] = []
    for row_number in range(header_row_index + 1, worksheet.max_row + 1):
        month_value = worksheet.cell(row=row_number, column=month_col).value
        month_key = _normalize_month_key(month_value)
        if not month_key:
            continue
        normalized = _normalize_history_row(
            {
                "기준년월": month_value,
                "매매일반거래가": read_value(row_number, SALE_GENERAL_KEYS),
                "전세일반거래가": read_value(row_number, JEONSE_GENERAL_KEYS),
                "매매상위평균가": read_value(row_number, SALE_UPPER_KEYS),
                "매매하위평균가": read_value(row_number, SALE_LOWER_KEYS),
                "전세상위평균가": read_value(row_number, JEONSE_UPPER_KEYS),
                "전세하위평균가": read_value(row_number, JEONSE_LOWER_KEYS),
            }
        )
        if normalized:
            rows.append(normalized)

    rows_by_month = {row["month_key"]: row for row in rows}
    return sorted(rows_by_month.values(), key=lambda row: row["month_key"], reverse=True)


def fetch_quote_history(
    complex_id: int,
    area_id: int,
    *,
    as_of_month: str | None = None,
    lookback_years: int = 10,
    force_refresh: bool = False,
) -> list[dict[str, Any]]:
    cached = valuation_db.get_price_history(complex_id, area_id)
    needed_months = max(lookback_years, 1) * 12
    if cached and not force_refresh and len(cached) >= needed_months:
        return cached

    years = _history_years(lookback_years, as_of_month)
    rows: list[dict[str, Any]] = []
    source = "json"

    try:
        rows = _request_quote_history_json(complex_id, area_id, years=years)
    except RuntimeError:
        rows = []

    if not rows:
        rows = _download_quote_history_excel(complex_id, area_id)
        source = "excel"

    if rows:
        valuation_db.replace_price_history(complex_id, area_id, rows, source=source)
        return valuation_db.get_price_history(complex_id, area_id)

    raise RuntimeError("KB 시세 이력을 가져오지 못했습니다.")


def _pick_base_type(complex_id: int, area_id: int) -> dict[str, Any]:
    for type_info in ensure_complex_types_cached(complex_id):
        if int(type_info["area_id"]) == int(area_id):
            return type_info
    raise ValueError("기준 단지 타입 정보를 찾지 못했습니다.")


def _match_peer_type(complex_id: int, target_exclusive_area: float | int | None) -> dict[str, Any]:
    rows = ensure_complex_types_cached(complex_id)
    if not rows:
        raise ValueError("비교 단지 타입 정보가 비어 있습니다.")
    if target_exclusive_area is None:
        return rows[0]

    best_row: dict[str, Any] | None = None
    best_score: tuple[float, int, int] | None = None
    for row in rows:
        area = row.get("exclusive_area")
        if area is None:
            continue
        diff = abs(float(area) - float(target_exclusive_area))
        score = (
            diff,
            abs(int(row.get("households") or 0)) * -1,
            abs(int(row.get("total_trade_count") or 0)) * -1,
        )
        if best_score is None or score < best_score:
            best_score = score
            best_row = row
    if best_row:
        return best_row
    return rows[0]


def _resolve_analysis_month(base_rows: list[dict[str, Any]], requested_month: str | None) -> str:
    if requested_month:
        requested_month = _normalize_month_key(requested_month)
        if requested_month:
            for row in base_rows:
                if row["month_key"] <= requested_month:
                    return row["month_key"]
    if not base_rows:
        raise ValueError("기준 단지의 시세 이력이 비어 있습니다.")
    return base_rows[0]["month_key"]


def _build_history_map(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {row["month_key"]: row for row in rows}


def _pick_common_month(common_months_desc: list[str], analysis_month: str) -> str | None:
    for month_key in common_months_desc:
        if month_key <= analysis_month:
            return month_key
    return common_months_desc[0] if common_months_desc else None


def _interpret_upside(value: float) -> str:
    if value >= 5:
        return "저평가 가능성 높음"
    if value >= 1:
        return "소폭 저평가"
    if value > -1:
        return "평균 수준"
    if value > -5:
        return "소폭 고평가"
    return "고평가 가능성"


def analyze_relative_value(
    *,
    base_complex_id: int,
    base_area_id: int,
    compare_complex_ids: list[int],
    as_of_month: str | None = None,
    lookback_years: int = 10,
    force_refresh: bool = False,
) -> dict[str, Any]:
    if not compare_complex_ids:
        raise ValueError("비교 단지를 1개 이상 선택해야 합니다.")

    base_complex = valuation_db.get_complex(base_complex_id)
    if not base_complex:
        raise ValueError("기준 단지 정보가 캐시에 없습니다. 먼저 지역 캐시를 동기화해 주세요.")

    base_type = _pick_base_type(base_complex_id, base_area_id)
    base_history = fetch_quote_history(
        base_complex_id,
        base_area_id,
        as_of_month=as_of_month,
        lookback_years=lookback_years,
        force_refresh=force_refresh,
    )
    analysis_month = _resolve_analysis_month(base_history, as_of_month)
    base_history_map = _build_history_map(base_history)
    base_current = base_history_map.get(analysis_month)
    if not base_current or not base_current.get("sale_general_price"):
        raise RuntimeError("기준 단지의 기준월 매매 일반평균가를 찾지 못했습니다.")

    lookback_month_count = max(lookback_years, 1) * 12
    comparisons: list[dict[str, Any]] = []

    for compare_complex_id in compare_complex_ids:
        if compare_complex_id == base_complex_id:
            continue

        compare_complex = valuation_db.get_complex(compare_complex_id)
        if not compare_complex:
            continue

        peer_type = _match_peer_type(compare_complex_id, base_type.get("exclusive_area"))
        peer_history = fetch_quote_history(
            compare_complex_id,
            int(peer_type["area_id"]),
            as_of_month=analysis_month,
            lookback_years=lookback_years,
            force_refresh=force_refresh,
        )
        peer_history_map = _build_history_map(peer_history)
        common_months_desc = sorted(
            set(base_history_map).intersection(peer_history_map),
            reverse=True,
        )
        current_month = _pick_common_month(common_months_desc, analysis_month)
        if not current_month:
            continue

        aligned_months = [month_key for month_key in common_months_desc if month_key <= current_month][:lookback_month_count]
        if not aligned_months:
            continue

        base_current_price = int(base_history_map[current_month]["sale_general_price"])
        peer_current_price = int(peer_history_map[current_month]["sale_general_price"])
        current_ratio = round((base_current_price / peer_current_price) * 100, 2) if peer_current_price else None
        ratio_series: list[dict[str, Any]] = []
        for month_key in aligned_months:
            base_price = base_history_map[month_key].get("sale_general_price")
            peer_price = peer_history_map[month_key].get("sale_general_price")
            if not base_price or not peer_price:
                continue
            ratio_series.append(
                {
                    "month_key": month_key,
                    "month_label": _month_label(month_key),
                    "base_price": int(base_price),
                    "compare_price": int(peer_price),
                    "ratio": round((int(base_price) / int(peer_price)) * 100, 2),
                }
            )

        if not ratio_series or current_ratio is None:
            continue

        average_ratio = round(mean(item["ratio"] for item in ratio_series), 2)
        upside = round(average_ratio - current_ratio, 2)
        comparisons.append(
            {
                "complex_id": compare_complex_id,
                "complex_name": compare_complex["complex_name"],
                "region": " ".join(filter(None, [compare_complex.get("sido_name"), compare_complex.get("sigungu_name")])),
                "current_month": current_month,
                "matched_area_id": peer_type["area_id"],
                "matched_area": peer_type.get("exclusive_area"),
                "matched_area_pyeong": peer_type.get("exclusive_area_pyeong"),
                "base_price": base_current_price,
                "compare_price": peer_current_price,
                "current_ratio": current_ratio,
                "avg_ratio_10y": average_ratio,
                "upside": upside,
                "label": _interpret_upside(upside),
                "history_months": len(ratio_series),
                "ratio_series": ratio_series,
            }
        )

    comparisons.sort(key=lambda item: item["upside"], reverse=True)

    return {
        "base": {
            "complex_id": base_complex_id,
            "complex_name": base_complex["complex_name"],
            "area_id": base_area_id,
            "area": base_type.get("exclusive_area"),
            "area_pyeong": base_type.get("exclusive_area_pyeong"),
            "analysis_month": analysis_month,
            "analysis_month_label": _month_label(analysis_month),
            "price": int(base_current["sale_general_price"]),
            "history_months": len(base_history),
        },
        "options": {
            "lookback_years": lookback_years,
            "force_refresh": force_refresh,
        },
        "comparisons": comparisons,
    }
