"""
KB부동산 주간 보도자료/docx + 시계열/xlsx 다운로드 및 분석
- 매매/전세 상하위 5개 지역 추출
"""
from __future__ import annotations

import os
import shutil
import subprocess
import tempfile
import zipfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote
from xml.etree import ElementTree as ET

import openpyxl
import requests

STATISTICS_API_URL = "https://api.kbland.kr/land-extra/statistics/reference"
FILE_DOWNLOAD_API_URL = "https://api.kbland.kr/land-extra/statistics/getfiledown"
DOWNLOAD_DIR = Path(__file__).parent / "downloads"
TIMEOUT_SEC = 30
TOP_N = 5

COMMON_HEADERS = {
    "accept": "application/json, text/plain, */*",
    "origin": "https://kbland.kr",
    "referer": "https://kbland.kr/webview.html#/main/statistics?blank=true",
    "user-agent": "Mozilla/5.0",
}

# 순위 집계에서 제외할 광역 단위 지역명
EXCLUDED_REGIONS = {
    "전국", "수도권", "6개광역시", "5개광역시", "강북14개구", "강남11개구", "기타지방",
}

# 시트 헤더에서 섹션 구분자로 사용되는 지역명 집합
SECTION_MARKERS = {
    "서울특별시", "6개광역시", "부산광역시", "대구광역시", "인천광역시", "광주광역시",
    "대전광역시", "울산광역시", "5개광역시", "수도권", "세종특별자치시", "경기도",
    "충청북도", "충청남도", "전라북도", "전북특별자치도", "전라남도", "경상북도",
    "경상남도", "제주도", "제주특별자치도", "기타지방",
}

CAPITAL_REGION_PREFIXES = ("서울특별시", "경기도", "인천광역시")


# ── 내부 헬퍼 ────────────────────────────────────────────────


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path

    stem = path.stem
    suffix = path.suffix
    index = 1
    while True:
        candidate = path.with_name(f"{stem} ({index}){suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def _list_existing_images(output_dir: Path) -> list[Path]:
    if not output_dir.exists():
        return []
    image_paths = sorted(path for path in output_dir.iterdir() if path.is_file())
    return _cleanup_displayable_images(image_paths)


def _detect_binary_image_suffix(binary: bytes, fallback_suffix: str) -> str:
    if binary.startswith(b"\x89PNG\r\n\x1a\n"):
        return ".png"
    if binary.startswith(b"\xff\xd8\xff"):
        return ".jpg"
    if binary.startswith((b"GIF87a", b"GIF89a")):
        return ".gif"
    if binary.startswith(b"BM"):
        return ".bmp"
    if binary.startswith(b"RIFF") and binary[8:12] == b"WEBP":
        return ".webp"
    if len(binary) > 44 and binary[:4] == b"\x01\x00\x00\x00" and binary[40:44] == b" EMF":
        return ".emf"
    return fallback_suffix or ".bin"


def _convert_vector_image_to_png(image_path: Path) -> Path | None:
    if image_path.suffix.lower() != ".emf":
        return None

    inkscape = shutil.which("inkscape")
    if inkscape:
        output_path = _unique_path(image_path.with_suffix(".png"))
        try:
            subprocess.run(
                [
                    inkscape,
                    str(image_path),
                    "--export-type=png",
                    f"--export-filename={output_path}",
                ],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=TIMEOUT_SEC,
            )
            if output_path.exists():
                return output_path
        except (subprocess.SubprocessError, OSError):
            pass

    office_bin = shutil.which("soffice") or shutil.which("libreoffice")
    if office_bin:
        temp_dir = Path(tempfile.mkdtemp(prefix="kb_emf_", dir=tempfile.gettempdir()))
        try:
            subprocess.run(
                [
                    office_bin,
                    "--headless",
                    "--convert-to",
                    "png",
                    "--outdir",
                    str(temp_dir),
                    str(image_path),
                ],
                check=True,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                timeout=TIMEOUT_SEC,
            )
            converted_temp = temp_dir / f"{image_path.stem}.png"
            if converted_temp.exists():
                output_path = _unique_path(image_path.with_suffix(".png"))
                converted_temp.replace(output_path)
                return output_path
        except (subprocess.SubprocessError, OSError):
            pass
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    return None


def _convert_extracted_images(image_paths: list[Path]) -> list[Path]:
    processed_paths = list(image_paths)
    for image_path in image_paths:
        converted_path = _convert_vector_image_to_png(image_path)
        if converted_path is not None:
            processed_paths.append(converted_path)
    return _cleanup_displayable_images(processed_paths)


def _cleanup_displayable_images(image_paths: list[Path]) -> list[Path]:
    displayable_suffixes = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
    existing_paths = [path for path in image_paths if path.exists()]

    for image_path in existing_paths:
        if image_path.suffix.lower() != ".emf":
            continue

        png_path = image_path.with_suffix(".png")
        if png_path.exists():
            image_path.unlink(missing_ok=True)

    cleaned_paths = [path for path in existing_paths if path.exists()]
    return sorted(path for path in cleaned_paths if path.suffix.lower() in displayable_suffixes)


def _extract_docx_images(docx_path: Path, output_dir: Path) -> list[Path]:
    existing = _list_existing_images(output_dir)
    if existing:
        return existing

    output_dir.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(docx_path) as zf:
        names = set(zf.namelist())
        if "word/document.xml" not in names or "word/_rels/document.xml.rels" not in names:
            return []

        rels_root = ET.fromstring(zf.read("word/_rels/document.xml.rels"))
        rels_map: dict[str, str] = {}
        for rel in rels_root:
            rel_id = rel.attrib.get("Id")
            target = rel.attrib.get("Target", "")
            if rel_id and target.startswith("media/"):
                rels_map[rel_id] = f"word/{target}"

        doc_root = ET.fromstring(zf.read("word/document.xml"))
        ns = {
            "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }

        ordered_internal_paths: list[str] = []
        for blip in doc_root.findall(".//a:blip", ns):
            rel_id = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
            internal_path = rels_map.get(rel_id or "")
            if internal_path and internal_path not in ordered_internal_paths:
                ordered_internal_paths.append(internal_path)

        if not ordered_internal_paths:
            ordered_internal_paths = sorted(name for name in names if name.startswith("word/media/"))

        extracted_paths: list[Path] = []
        for index, internal_path in enumerate(ordered_internal_paths, start=1):
            if internal_path not in names:
                continue

            binary = zf.read(internal_path)
            suffix = _detect_binary_image_suffix(binary, Path(internal_path).suffix)
            save_path = _unique_path(output_dir / f"image_{index:02d}{suffix}")
            save_path.write_bytes(binary)
            extracted_paths.append(save_path)

    return _convert_extracted_images(extracted_paths)

def _this_month_range(today: date | None = None) -> tuple[str, str]:
    today = today or date.today()
    month_start = today.replace(day=1)
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    month_end = next_month - timedelta(days=1)
    return month_start.strftime("%Y-%m-%d"), month_end.strftime("%Y-%m-%d")


def _fetch_statistics_meta() -> dict[str, Any]:
    """KB 통계자료실 API 응답(dataBody.data) 반환."""
    start_date, end_date = _this_month_range()
    params = {
        "주월간구분": 0,
        "기준년월시작일": start_date,
        "기준년월종료일": end_date,
    }
    resp = requests.get(STATISTICS_API_URL, params=params, headers=COMMON_HEADERS, timeout=TIMEOUT_SEC)
    resp.raise_for_status()
    payload = resp.json()

    header = payload.get("dataHeader", {})
    if header.get("resultCode") != "10000":
        raise RuntimeError(f"KB API 오류: {header}")

    data = payload.get("dataBody", {}).get("data")
    if not data:
        raise RuntimeError("KB API 응답 데이터가 비어 있습니다.")
    return data


def _pick_week_files(data: dict[str, Any]) -> tuple[str, dict[str, Any], dict[str, Any]]:
    """
    통계 메타에서 최신 주차 보도자료(docx)와 시계열(xlsx) 추출.
    반환: (통계최신일, price_doc_file, series_file)

    data 구조 (원본 참고):
      data["통계최신일"] = "2026-03-14"
      data["가격동향"] = [{"2026-03-14": [{...}, {...}]}]
      data["시계열"] = [{"파일구분": "3", "파일경로": ..., "파일명": ..., "원본파일명": ...}, ...]
    """
    latest_date = data.get("통계최신일", "")
    if not latest_date:
        raise RuntimeError("통계최신일을 찾을 수 없습니다.")

    price_trend_list = data.get("가격동향", [])
    latest_price_files: list[dict[str, Any]] = []
    for bucket in price_trend_list:
        if latest_date in bucket:
            latest_price_files = bucket[latest_date]
            break
    if not latest_price_files:
        raise RuntimeError(f"가격동향에서 최신일({latest_date}) 데이터를 찾지 못했습니다.")

    price_doc = next((f for f in latest_price_files if str(f.get("파일구분")) == "1"), None)
    if not price_doc:
        raise RuntimeError("가격동향 보도자료 파일(파일구분 1)을 찾지 못했습니다.")

    series_list = data.get("시계열", [])
    series_file = next((f for f in series_list if str(f.get("파일구분")) == "3"), None)
    if not series_file:
        raise RuntimeError("시계열 파일(파일구분 3)을 찾지 못했습니다.")
    return latest_date, price_doc, series_file


def _build_download_url(file_info: dict[str, Any]) -> str:
    path_part = f"{file_info['파일경로']}/{file_info['파일명']}"
    encoded_path = quote(path_part, safe="/:")
    encoded_name = quote(file_info["원본파일명"], safe="")
    return f"{FILE_DOWNLOAD_API_URL}?urlpath={encoded_path}&filename={encoded_name}"


def _download_file(file_info: dict[str, Any], dest_dir: Path) -> Path:
    filename = file_info["원본파일명"].replace("/", "_").strip()
    save_path = dest_dir / filename
    # 이미 다운로드된 경우 재사용
    if save_path.exists():
        return save_path

    url = _build_download_url(file_info)
    resp = requests.get(url, headers=COMMON_HEADERS, timeout=TIMEOUT_SEC, stream=True)
    resp.raise_for_status()

    with open(save_path, "wb") as f:
        for chunk in resp.iter_content(chunk_size=65536):
            if chunk:
                f.write(chunk)
    return save_path


def _find_latest_two_rows(ws) -> tuple[list[Any], list[Any]]:
    """시트에서 날짜 기준 마지막 2개 데이터 행 반환."""
    prev_row: list[Any] | None = None
    last_row: list[Any] | None = None
    for row in ws.iter_rows(min_row=1, values_only=True):
        if isinstance(row[0], datetime):
            prev_row = last_row
            last_row = list(row)
    if prev_row is None or last_row is None:
        raise RuntimeError(f"시트 '{ws.title}'에서 데이터 행을 2개 이상 찾지 못했습니다.")
    return prev_row, last_row


def _parse_change_sheet(ws) -> tuple[str, list[dict[str, Any]]]:
    """
    매매증감/전세증감 시트 파싱.
    반환: (기준일 문자열, 지역별 변동 리스트)
    각 항목: {"region": str, "current": float, "delta": float}
    """
    # 2행이 헤더(지역명)
    header = list(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)))
    prev_row, last_row = _find_latest_two_rows(ws)

    latest_date = last_row[0].strftime("%Y-%m-%d") if isinstance(last_row[0], datetime) else str(last_row[0])
    current_section = ""
    current_city = ""
    records: list[dict[str, Any]] = []

    for idx, raw in enumerate(header):
        if idx == 0 or not isinstance(raw, str):
            continue
        region = raw.strip()
        if not region:
            continue

        # 세종특별자치시 처리
        if region.startswith("세종특별자치"):
            region = "세종특별자치시"

        is_section = region in SECTION_MARKERS or region.startswith("세종특별자치")
        is_city = region.endswith("시") and not region.endswith(("특별시", "광역시", "특별자치시"))

        if is_section:
            current_section = region
            current_city = ""
        elif is_city:
            current_city = region

        curr = last_row[idx] if idx < len(last_row) else None
        prev = prev_row[idx] if idx < len(prev_row) else None
        if not isinstance(curr, (int, float)) or not isinstance(prev, (int, float)):
            continue

        if region in EXCLUDED_REGIONS:
            continue

        # 표시명 조합: 섹션 + 시 + 지역구
        display_parts: list[str] = []
        if current_section and region != current_section:
            display_parts.append(current_section)
        if current_city and region not in {current_section, current_city}:
            display_parts.append(current_city)
        display_parts.append(region)

        records.append({
            "region": " ".join(display_parts),
            "current": float(curr),
            "delta": float(curr - prev),
        })

    return latest_date, records


def _top_bottom(records: list[dict[str, Any]], n: int = TOP_N) -> dict[str, list[dict]]:
    """current 기준 상위 n / 하위 n 추출."""
    sorted_desc = sorted(records, key=lambda x: x["current"], reverse=True)
    sorted_asc = sorted(records, key=lambda x: x["current"])
    return {
        f"top{n}": [
            {
                "region": r["region"],
                "current": round(r["current"], 3),
                "delta": round(r["delta"], 3),
            }
            for r in sorted_desc[:n]
        ],
        f"bottom{n}": [
            {
                "region": r["region"],
                "current": round(r["current"], 3),
                "delta": round(r["delta"], 3),
            }
            for r in sorted_asc[:n]
        ],
    }


def _select_ranked_regions(
    records: list[dict[str, Any]],
    n: int,
    *,
    reverse: bool,
) -> list[dict[str, Any]]:
    sorted_records = sorted(records, key=lambda x: x["current"], reverse=reverse)
    return [
        {
            "region": r["region"],
            "current": round(r["current"], 3),
            "delta": round(r["delta"], 3),
        }
        for r in sorted_records[:n]
    ]


def _is_capital_region(region_name: str) -> bool:
    return region_name.startswith(CAPITAL_REGION_PREFIXES)


def _split_capital_regions(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    capital_records: list[dict[str, Any]] = []
    non_capital_records: list[dict[str, Any]] = []

    for record in records:
        if _is_capital_region(record["region"]):
            capital_records.append(record)
        else:
            non_capital_records.append(record)

    return capital_records, non_capital_records


def extract_content_regions(data: dict[str, Any]) -> dict[str, Any]:
    """
    콘텐츠용 지역 8개 버킷 추출.

    - 수도권 매매 상승 상위 5 / 하위 5
    - 수도권 전세 상승 상위 5 / 하위 5
    - 비수도권 매매 상승 상위 5 / 하위 5
    - 비수도권 전세 상승 상위 5 / 하위 5
    """
    sale_capital, sale_non_capital = _split_capital_regions(data["sale"])
    rent_capital, rent_non_capital = _split_capital_regions(data["rent"])

    return {
        "capital_sale_top5": _select_ranked_regions(sale_capital, TOP_N, reverse=True),
        "capital_sale_bottom5": _select_ranked_regions(sale_capital, TOP_N, reverse=False),
        "capital_rent_top5": _select_ranked_regions(rent_capital, TOP_N, reverse=True),
        "capital_rent_bottom5": _select_ranked_regions(rent_capital, TOP_N, reverse=False),
        "non_capital_sale_top5": _select_ranked_regions(sale_non_capital, TOP_N, reverse=True),
        "non_capital_sale_bottom5": _select_ranked_regions(sale_non_capital, TOP_N, reverse=False),
        "non_capital_rent_top5": _select_ranked_regions(rent_non_capital, TOP_N, reverse=True),
        "non_capital_rent_bottom5": _select_ranked_regions(rent_non_capital, TOP_N, reverse=False),
    }


# ── 공개 API ────────────────────────────────────────────────

def download_kb_files() -> tuple[dict[str, Any], str]:
    """
    KB부동산 주간 보도자료(docx), 시계열(xlsx), 보도자료 이미지 자동 다운로드/추출.
    반환:
      (
        {
          "docx": 저장된 보도자료 파일 경로,
          "xlsx": 저장된 시계열 파일 경로,
          "image_dir": 추출 이미지 폴더 경로,
          "images": 추출 이미지 파일 경로 리스트,
        },
        통계최신일 "YYYY-MM-DD",
      )
    """
    DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    data = _fetch_statistics_meta()
    latest_date, docx_info, xlsx_info = _pick_week_files(data)
    downloaded = {
        "docx": _download_file(docx_info, DOWNLOAD_DIR),
        "xlsx": _download_file(xlsx_info, DOWNLOAD_DIR),
    }

    image_dir = DOWNLOAD_DIR / f"{downloaded['docx'].stem}_images"
    downloaded["image_dir"] = image_dir
    downloaded["images"] = _extract_docx_images(downloaded["docx"], image_dir)
    return downloaded, latest_date


def download_kb_excel() -> tuple[Path, str]:
    """
    시계열 xlsx 경로만 필요할 때 사용하는 호환용 래퍼.
    보도자료 docx도 함께 다운로드된다.
    """
    downloaded, latest_date = download_kb_files()
    return downloaded["xlsx"], latest_date


def parse_excel(file_path: str | Path) -> dict[str, Any]:
    """
    시계열 xlsx 파싱.
    반환:
      {
        "latest_date": "2026-03-14",
        "sale": [{"region": ..., "current": ..., "delta": ...}, ...],
        "rent": [{"region": ..., "current": ..., "delta": ...}, ...],
      }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet_names = wb.sheetnames
        if len(sheet_names) < 3:
            raise RuntimeError("시트가 3개 미만입니다. 파일을 확인하세요.")

        sale_date, sale_records = _parse_change_sheet(wb[sheet_names[1]])
        rent_date, rent_records = _parse_change_sheet(wb[sheet_names[2]])
    finally:
        wb.close()

    latest_date = sale_date or rent_date
    return {
        "latest_date": latest_date,
        "sale": sale_records,
        "rent": rent_records,
    }


def extract_top_bottom(data: dict[str, Any], n: int = TOP_N) -> dict[str, Any]:
    """
    parse_excel() 결과를 받아 매매/전세 상하위 n개 추출.
    반환 형식:
      {
        "latest_date": "2026-03-14",
        "sale": {"top5": [...], "bottom5": [...]},
        "rent": {"top5": [...], "bottom5": [...]},
        "content_regions": {
          "capital_sale_top5": [...],
          "capital_sale_bottom5": [...],
          "capital_rent_top5": [...],
          "capital_rent_bottom5": [...],
          "non_capital_sale_top5": [...],
          "non_capital_sale_bottom5": [...],
          "non_capital_rent_top5": [...],
          "non_capital_rent_bottom5": [...],
        }
      }
    """
    return {
        "latest_date": data["latest_date"],
        "sale": _top_bottom(data["sale"], n),
        "rent": _top_bottom(data["rent"], n),
        "content_regions": extract_content_regions(data),
    }


def run_analysis() -> dict[str, Any]:
    """
    전체 분석 파이프라인 실행.
    1. 보도자료 docx + 시계열 xlsx 다운로드
    2. xlsx 시트 파싱
    3. 상하위 5개 추출
    반환 형식:
      {
        "latest_date": "2026-03-14",
        "sale": {
          "top5": [{"region": "강동구", "current": 102.314, "delta": 0.113}, ...],
          "bottom5": [...]
        },
        "rent": {
          "top5": [...],
          "bottom5": [...]
        },
        "content_regions": {
          "capital_sale_top5": [...],
          "capital_sale_bottom5": [...],
          "capital_rent_top5": [...],
          "capital_rent_bottom5": [...],
          "non_capital_sale_top5": [...],
          "non_capital_sale_bottom5": [...],
          "non_capital_rent_top5": [...],
          "non_capital_rent_bottom5": [...],
        },
        "source_files": {
          "docx": "...",
          "xlsx": "...",
          "image_dir": "..."
        },
        "report_images": ["...", "..."],
      }
    """
    downloaded_files, latest_date = download_kb_files()
    raw = parse_excel(downloaded_files["xlsx"])
    raw["latest_date"] = latest_date  # API 기준일 우선 사용
    result = extract_top_bottom(raw)
    result["source_files"] = {
        "docx": str(downloaded_files["docx"]),
        "xlsx": str(downloaded_files["xlsx"]),
        "image_dir": str(downloaded_files["image_dir"]),
    }
    result["report_images"] = [str(path) for path in downloaded_files.get("images", [])]
    return result
